"""
Fetches monitored-universe constituent lists.

- KOSPI 200: Naver Finance constituent page
- KOSDAQ 150: PLUS ETF holdings export (filters out non-equity rows)
"""

import io
import os
import time
import logging
import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

NAVER_INDEX_CONSTITUENTS_URL = "https://finance.naver.com/sise/entryJongmok.naver"
PLUS_KOSDAQ150_PRODUCT_ID = "006318"
PLUS_KOSDAQ150_EXPORT_URL = "https://www.plusetf.co.kr/excel/product/pdf"
DEFAULT_USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
HEADERS = {
    "User-Agent": os.environ.get("SCRAPER_USER_AGENT", DEFAULT_USER_AGENT),
    "Accept-Language": "ko-KR,ko;q=0.9",
}
logger = logging.getLogger(__name__)


def _request_with_retry(client: httpx.Client, url: str, **kwargs) -> httpx.Response:
    last_exc = None
    for attempt in range(3):
        try:
            resp = client.get(url, **kwargs)
            resp.raise_for_status()
            return resp
        except httpx.HTTPError as exc:
            last_exc = exc
            if attempt == 2:
                raise
            delay = 2 ** attempt
            logger.warning("KRX fetch failed (%s). Retrying in %ss", exc, delay)
            time.sleep(delay)
    raise last_exc


def _fetch_index_constituents(index_code: str, total_pages: int) -> list[dict]:
    results = []
    with httpx.Client(timeout=30, headers=HEADERS) as client:
        for page in range(1, total_pages + 1):
            resp = _request_with_retry(
                client,
                NAVER_INDEX_CONSTITUENTS_URL,
                params={"code": index_code, "page": str(page)},
            )

            soup = BeautifulSoup(resp.content.decode("euc-kr", errors="replace"), "lxml")
            links = soup.select('a[href*="/item/main.naver?code="]')

            for anchor in links:
                ticker = anchor["href"].split("code=")[-1].strip()
                company = anchor.get_text(strip=True)
                if ticker and company:
                    results.append({"ticker": ticker, "company": company})

    seen = set()
    unique = []
    for row in results:
        if row["ticker"] not in seen:
            seen.add(row["ticker"])
            unique.append(row)

    return unique


def fetch_kospi200() -> list[dict]:
    """
    Returns list of {ticker, company} for current KOSPI 200 constituents.
    """
    return _fetch_index_constituents("KPI200", total_pages=20)


def fetch_kosdaq150() -> list[dict]:
    """
    Returns list of {ticker, company} for current KOSDAQ 150 constituents.
    """
    today = time.strftime("%Y%m%d")
    with httpx.Client(timeout=30, headers=HEADERS, follow_redirects=True) as client:
        resp = _request_with_retry(
            client,
            PLUS_KOSDAQ150_EXPORT_URL,
            params={
                "n": PLUS_KOSDAQ150_PRODUCT_ID,
                "d": today,
                "title": "PLUS 코스닥150",
            },
        )

    workbook = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    results = []
    for row in worksheet.iter_rows(values_only=True):
        if not row or len(row) < 3:
            continue
        raw_ticker = row[1]
        raw_company = row[2]
        if raw_ticker in (None, "종목코드"):
            continue

        ticker = str(raw_ticker).strip()
        company = str(raw_company).strip() if raw_company is not None else ""

        # ETF holdings export contains one cash row such as `원화예금` with a long code.
        # Keep 6-character alphanumeric short codes because some valid constituents
        # can use non-digit symbols (for example, preferred/newly listed variants).
        if len(ticker) == 6 and ticker.isalnum() and company:
            results.append({"ticker": ticker, "company": company})

    seen = set()
    unique = []
    for row in results:
        if row["ticker"] not in seen:
            seen.add(row["ticker"])
            unique.append(row)

    return unique


if __name__ == "__main__":
    kospi200 = fetch_kospi200()
    kosdaq150 = fetch_kosdaq150()
    print(f"Fetched {len(kospi200)} KOSPI 200 constituents")
    print(f"Fetched {len(kosdaq150)} KOSDAQ 150 constituents")
